"""Phase 7: Translation failure scenario tests.

Covers:
- API timeout
- Invalid API key (401)
- Rate limit (429)
- JSON parsing failure
- Array length mismatch
- Partial batch failure with successful segments preserved
"""

import json
from unittest.mock import MagicMock, patch

import pytest

from yt_excel.translator import (
    Batch,
    TranslationResult,
    call_translation_api,
    parse_translation_response,
    translate_batch_with_retry,
    translate_segments,
    validate_translations,
)
from yt_excel.config import TranslationConfig
from yt_excel.vtt import Segment


def _make_segments(count: int) -> list[Segment]:
    """Create a list of test segments."""
    return [
        Segment(
            index=i + 1,
            start=f"00:00:{i:02d}.000",
            end=f"00:00:{i + 3:02d}.000",
            english=f"Sentence number {i + 1}.",
        )
        for i in range(count)
    ]


def _make_batch(count: int) -> Batch:
    """Create a test batch with the given number of translate segments."""
    segments = _make_segments(count)
    return Batch(
        translate_segments=segments,
        context_before=[],
        context_after=[],
    )


class TestAPITimeout:
    """Tests for API timeout handling."""

    def test_timeout_triggers_retry(self):
        """APITimeoutError triggers retry logic."""
        from openai import APITimeoutError

        batch = _make_batch(3)
        client = MagicMock()

        # First call times out, second succeeds
        good_response = MagicMock()
        good_response.choices = [MagicMock()]
        good_response.choices[0].message.content = json.dumps(
            {"translations": ["번역1", "번역2", "번역3"]}
        )

        client.chat.completions.create.side_effect = [
            APITimeoutError(request=MagicMock()),
            good_response,
        ]

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["번역1", "번역2", "번역3"]
        assert client.chat.completions.create.call_count == 2

    def test_all_timeouts_return_empty(self):
        """All attempts timing out returns empty translations."""
        from openai import APITimeoutError

        batch = _make_batch(3)
        client = MagicMock()
        client.chat.completions.create.side_effect = APITimeoutError(
            request=MagicMock()
        )

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["", "", ""]
        assert client.chat.completions.create.call_count == 3


class TestInvalidAPIKey:
    """Tests for invalid API key (401 Unauthorized)."""

    def test_auth_error_not_retried_by_batch(self):
        """AuthenticationError is NOT in the retry list, so it propagates.

        Note: translate_batch_with_retry only catches RateLimitError,
        APIConnectionError, APITimeoutError, and ValueError. Auth errors
        will propagate up to the caller.
        """
        from openai import AuthenticationError

        batch = _make_batch(2)
        client = MagicMock()

        # Create a proper AuthenticationError
        mock_response = MagicMock()
        mock_response.status_code = 401
        mock_response.json.return_value = {"error": {"message": "Invalid API key"}}
        client.chat.completions.create.side_effect = AuthenticationError(
            message="Invalid API key",
            response=mock_response,
            body={"error": {"message": "Invalid API key"}},
        )

        with patch("yt_excel.translator.time.sleep"):
            with pytest.raises(AuthenticationError):
                translate_batch_with_retry(
                    client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
                )


class TestRateLimit:
    """Tests for rate limiting (429)."""

    def test_429_triggers_retry_with_backoff(self):
        """RateLimitError triggers retry with exponential backoff."""
        from openai import RateLimitError

        batch = _make_batch(2)
        client = MagicMock()

        # First call gets rate limited, second succeeds
        mock_response = MagicMock()
        mock_response.status_code = 429
        mock_response.headers = {"retry-after": "1"}
        mock_response.json.return_value = {"error": {"message": "Rate limit"}}

        good_response = MagicMock()
        good_response.choices = [MagicMock()]
        good_response.choices[0].message.content = json.dumps(
            {"translations": ["번역1", "번역2"]}
        )

        client.chat.completions.create.side_effect = [
            RateLimitError(
                message="Rate limit exceeded",
                response=mock_response,
                body={"error": {"message": "Rate limit"}},
            ),
            good_response,
        ]

        with patch("yt_excel.translator.time.sleep") as mock_sleep:
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["번역1", "번역2"]

    def test_429_exhausted_returns_empty(self):
        """All 429 responses exhaust retries and return empty."""
        from openai import RateLimitError

        batch = _make_batch(2)
        client = MagicMock()

        mock_response = MagicMock()
        mock_response.status_code = 429
        mock_response.headers = {}
        mock_response.json.return_value = {"error": {"message": "Rate limit"}}

        client.chat.completions.create.side_effect = RateLimitError(
            message="Rate limit exceeded",
            response=mock_response,
            body={"error": {"message": "Rate limit"}},
        )

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["", ""]


class TestJSONParsingFailure:
    """Tests for JSON parsing failures in API responses."""

    def test_invalid_json_triggers_retry(self):
        """Invalid JSON response triggers retry."""
        batch = _make_batch(2)
        client = MagicMock()

        # First response: invalid JSON; second: valid
        bad_response = MagicMock()
        bad_response.choices = [MagicMock()]
        bad_response.choices[0].message.content = "This is not JSON at all"

        good_response = MagicMock()
        good_response.choices = [MagicMock()]
        good_response.choices[0].message.content = json.dumps(
            {"translations": ["번역1", "번역2"]}
        )

        client.chat.completions.create.side_effect = [bad_response, good_response]

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["번역1", "번역2"]

    def test_all_invalid_json_returns_empty(self):
        """All invalid JSON responses return empty translations."""
        batch = _make_batch(2)
        client = MagicMock()

        bad_response = MagicMock()
        bad_response.choices = [MagicMock()]
        bad_response.choices[0].message.content = "NOT JSON"

        client.chat.completions.create.return_value = bad_response

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["", ""]

    def test_markdown_wrapped_json_parsed(self):
        """JSON wrapped in markdown code block is handled."""
        raw = '```json\n{"translations": ["안녕", "세상"]}\n```'
        result = parse_translation_response(raw, 2)
        assert result == ["안녕", "세상"]

    def test_response_with_extra_fields_parsed(self):
        """Response with extra fields still extracts translations."""
        raw = json.dumps({
            "translations": ["번역1", "번역2"],
            "confidence": [0.95, 0.88],
            "notes": "some extra info",
        })
        result = parse_translation_response(raw, 2)
        assert result == ["번역1", "번역2"]

    def test_empty_response_triggers_error(self):
        """Empty string response triggers ValueError."""
        with pytest.raises(ValueError):
            parse_translation_response("", 2)

    def test_null_content_handled(self):
        """None content from API treated as empty."""
        batch = _make_batch(2)
        client = MagicMock()

        response = MagicMock()
        response.choices = [MagicMock()]
        response.choices[0].message.content = None

        client.chat.completions.create.return_value = response

        # call_translation_api returns "" for None content
        # which triggers ValueError in parse_translation_response
        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["", ""]


class TestArrayLengthMismatch:
    """Tests for translation array length mismatches."""

    def test_excess_translations_truncated(self):
        """More translations than expected: truncate to N."""
        translations = ["번역1", "번역2", "번역3", "추가1", "추가2"]
        result = validate_translations(translations, 3)
        assert result == ["번역1", "번역2", "번역3"]

    def test_fewer_translations_raises_value_error(self):
        """Fewer translations than expected raises ValueError."""
        translations = ["번역1"]
        with pytest.raises(ValueError, match="expected 3"):
            validate_translations(translations, 3)

    def test_empty_translations_raises_value_error(self):
        """Empty translations array raises ValueError."""
        with pytest.raises(ValueError, match="expected 3"):
            validate_translations([], 3)

    def test_excess_in_batch_retry_succeeds(self):
        """Excess translations are handled gracefully during batch retry."""
        batch = _make_batch(2)
        client = MagicMock()

        response = MagicMock()
        response.choices = [MagicMock()]
        response.choices[0].message.content = json.dumps(
            {"translations": ["번역1", "번역2", "번역3"]}  # 3 instead of 2
        )

        client.chat.completions.create.return_value = response

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["번역1", "번역2"]

    def test_shortage_retry_then_success(self):
        """Shortage triggers retry, next attempt succeeds."""
        batch = _make_batch(3)
        client = MagicMock()

        short_response = MagicMock()
        short_response.choices = [MagicMock()]
        short_response.choices[0].message.content = json.dumps(
            {"translations": ["번역1"]}  # Only 1, expected 3
        )

        good_response = MagicMock()
        good_response.choices = [MagicMock()]
        good_response.choices[0].message.content = json.dumps(
            {"translations": ["번역1", "번역2", "번역3"]}
        )

        client.chat.completions.create.side_effect = [short_response, good_response]

        with patch("yt_excel.translator.time.sleep"):
            result = translate_batch_with_retry(
                client, batch, "gpt-5-nano", max_retries=3, request_interval_ms=0,
            )

        assert result == ["번역1", "번역2", "번역3"]


class TestPartialBatchFailure:
    """Tests for partial translation failure (some batches fail, others succeed)."""

    def test_partial_failure_preserves_successful_segments(self):
        """When one batch fails, other batches' translations are preserved."""
        segments = _make_segments(15)  # 2 batches: 10 + 5
        config = TranslationConfig(
            model="gpt-5-nano",
            batch_size=10,
            context_before=0,
            context_after=0,
            max_retries=1,
            request_interval_ms=0,
        )

        client = MagicMock()

        # First batch succeeds
        good_response = MagicMock()
        good_response.choices = [MagicMock()]
        good_response.choices[0].message.content = json.dumps(
            {"translations": [f"번역{i}" for i in range(1, 11)]}
        )

        # Second batch fails (invalid JSON)
        bad_response = MagicMock()
        bad_response.choices = [MagicMock()]
        bad_response.choices[0].message.content = "INVALID"

        client.chat.completions.create.side_effect = [good_response, bad_response]

        with patch("yt_excel.translator.time.sleep"):
            result = translate_segments(client, segments, config)

        assert result.success_count == 10
        assert result.failed_count == 5
        assert len(result.segments) == 15

        # First 10 have translations
        for i in range(10):
            assert result.segments[i].korean == f"번역{i + 1}"

        # Last 5 have empty Korean
        for i in range(10, 15):
            assert result.segments[i].korean == ""

    def test_all_batches_fail_all_empty(self):
        """When all batches fail, all segments have empty Korean."""
        segments = _make_segments(5)
        config = TranslationConfig(
            model="gpt-5-nano",
            batch_size=10,
            context_before=0,
            context_after=0,
            max_retries=1,
            request_interval_ms=0,
        )

        client = MagicMock()
        bad_response = MagicMock()
        bad_response.choices = [MagicMock()]
        bad_response.choices[0].message.content = "NOT JSON"
        client.chat.completions.create.return_value = bad_response

        with patch("yt_excel.translator.time.sleep"):
            result = translate_segments(client, segments, config)

        assert result.success_count == 0
        assert result.failed_count == 5
        for seg in result.segments:
            assert seg.korean == ""

    def test_failed_segments_written_to_excel(self, tmp_path):
        """Failed translation segments (empty korean) are saved to Excel."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        segments = [
            Segment(index=1, start="00:00:01.000", end="00:00:04.000",
                    english="Success", korean="성공"),
            Segment(index=2, start="00:00:05.000", end="00:00:08.000",
                    english="Failed", korean=""),  # Empty = failed
            Segment(index=3, start="00:00:09.000", end="00:00:12.000",
                    english="Also success", korean="역시 성공"),
        ]

        wb = openpyxl.load_workbook(str(master_path))
        write_data_sheet(wb, "Test", segments)
        wb.save(str(master_path))

        # Verify data is written correctly
        wb2 = openpyxl.load_workbook(str(master_path))
        ws = wb2["Test"]
        assert ws.cell(row=2, column=5).value == "성공"
        assert not ws.cell(row=3, column=5).value  # Empty/None = failed
        assert ws.cell(row=4, column=5).value == "역시 성공"


# Need this import for the Excel test in TestPartialBatchFailure
import openpyxl
from yt_excel.excel import initialize_workbook, write_data_sheet
