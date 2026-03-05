"""Phase 7: End-to-end pipeline tests with mocked external services.

Covers:
- Full pipeline run (URL -> Excel) with mocked yt-dlp and OpenAI
- Caption-less video handling
- Auto-caption-only video handling
- en-US / en-GB language code handling
- Duplicate video detection
- Dry-run mode
"""

import json
import os
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from yt_excel.cli import Output, _run_pipeline, build_parser
from yt_excel.config import AppConfig
from yt_excel.excel import METADATA_SHEET, STUDY_LOG_SHEET, initialize_workbook
from yt_excel.youtube import (
    AutoCaptionOnlyError,
    CaptionNotFoundError,
    CaptionInfo,
    VideoMeta,
)


# --- Fixtures ---

SAMPLE_VTT = """\
WEBVTT

00:00:01.000 --> 00:00:04.500
Hello and welcome to this video.

00:00:05.000 --> 00:00:08.200
Today we are going to learn about DNA.

00:00:09.000 --> 00:00:12.800
DNA is the blueprint of life.

00:00:13.500 --> 00:00:17.000
It contains all the instructions for building an organism.

00:00:18.000 --> 00:00:21.500
Thank you for watching.
"""


def _make_args(url: str, tmp_path: Path, dry_run: bool = False):
    """Create mock CLI args."""
    parser = build_parser()
    args_list = [url, "--master", str(tmp_path / "Master.xlsx")]
    if dry_run:
        args_list.append("--dry-run")
    return parser.parse_args(args_list)


def _make_config(tmp_path: Path) -> AppConfig:
    """Create test config."""
    config = AppConfig()
    config.file.master_path = str(tmp_path / "Master.xlsx")
    config.translation.max_retries = 1
    config.translation.request_interval_ms = 0
    return config


def _mock_translation_response(count: int) -> MagicMock:
    """Create a mock successful translation response."""
    response = MagicMock()
    response.choices = [MagicMock()]
    response.choices[0].message.content = json.dumps(
        {"translations": [f"한국어 번역 {i}" for i in range(1, count + 1)]}
    )
    return response


class TestFullPipelineE2E:
    """Full pipeline end-to-end tests with mocked external services."""

    @patch("yt_excel.cli.download_captions")
    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    @patch("yt_excel.cli.validate_api_key")
    @patch("yt_excel.translator.time.sleep")
    def test_successful_pipeline(
        self,
        mock_sleep,
        mock_api_key,
        mock_fetch_meta,
        mock_list_captions,
        mock_download,
        tmp_path,
    ):
        """Full pipeline succeeds with mocked services."""
        mock_api_key.return_value = "sk-test"
        mock_fetch_meta.return_value = VideoMeta(
            video_id="dQw4w9WgXcQ",
            title="How DNA Works",
            channel="TED-Ed",
            duration="00:04:52",
        )
        mock_list_captions.return_value = CaptionInfo(
            lang_code="en", caption_type="manual", available_codes=["en"]
        )
        mock_download.return_value = SAMPLE_VTT

        args = _make_args(
            "https://www.youtube.com/watch?v=dQw4w9WgXcQ", tmp_path
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        # Mock OpenAI client
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = _mock_translation_response(5)

        with patch("yt_excel.cli.create_client", return_value=mock_client):
            with patch("yt_excel.cli.detect_font", return_value="Malgun Gothic"):
                _run_pipeline(args, config, out, 0.0)

        # Verify Excel was created
        master_path = tmp_path / "Master.xlsx"
        assert master_path.exists()

        wb = openpyxl.load_workbook(str(master_path))
        assert "How DNA Works" in wb.sheetnames
        assert METADATA_SHEET in wb.sheetnames
        assert STUDY_LOG_SHEET in wb.sheetnames

        # Verify data sheet
        ws = wb["How DNA Works"]
        assert ws.cell(row=1, column=1).value == "Index"
        assert ws.cell(row=2, column=4).value == "Hello and welcome to this video."

        # Verify metadata
        ws_meta = wb[METADATA_SHEET]
        assert ws_meta.cell(row=2, column=1).value == "dQw4w9WgXcQ"

        # Verify study log
        ws_log = wb[STUDY_LOG_SHEET]
        assert ws_log.cell(row=2, column=3).value == "How DNA Works"
        assert ws_log.cell(row=2, column=6).value == "Not Started"


class TestNoCaptionsVideo:
    """Tests for videos with no captions."""

    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    @patch("yt_excel.cli.validate_api_key")
    def test_no_captions_exits_with_error(
        self, mock_api_key, mock_fetch_meta, mock_list_captions, tmp_path
    ):
        """Video with no captions exits with error message."""
        mock_api_key.return_value = "sk-test"
        mock_fetch_meta.return_value = VideoMeta(
            video_id="nocaption12",
            title="No Captions Video",
            channel="TestChannel",
            duration="00:05:00",
        )
        mock_list_captions.side_effect = CaptionNotFoundError(
            "No English captions found for this video."
        )

        args = _make_args(
            "https://www.youtube.com/watch?v=nocaption12", tmp_path
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        with pytest.raises(SystemExit) as exc_info:
            _run_pipeline(args, config, out, 0.0)
        assert exc_info.value.code == 1


class TestAutoCaptionOnlyVideo:
    """Tests for videos with only auto-generated captions."""

    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    @patch("yt_excel.cli.validate_api_key")
    def test_auto_caption_only_exits_with_warning(
        self, mock_api_key, mock_fetch_meta, mock_list_captions, tmp_path
    ):
        """Video with only auto captions exits with warning."""
        mock_api_key.return_value = "sk-test"
        mock_fetch_meta.return_value = VideoMeta(
            video_id="autocap1234",
            title="Auto Caption Video",
            channel="TestChannel",
            duration="00:03:00",
        )
        mock_list_captions.side_effect = AutoCaptionOnlyError(
            "This video only has auto-generated English captions."
        )

        args = _make_args(
            "https://www.youtube.com/watch?v=autocap1234", tmp_path
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        with pytest.raises(SystemExit) as exc_info:
            _run_pipeline(args, config, out, 0.0)
        assert exc_info.value.code == 1


class TestLanguageCodeVariants:
    """Tests for en-US, en-GB, and other English variant codes."""

    def test_en_us_recognized(self):
        """en-US is recognized as English by _find_english_code."""
        from yt_excel.youtube import _find_english_code

        lang_dict = {"en-US": [{"ext": "vtt"}], "fr": [{"ext": "vtt"}]}
        result = _find_english_code(lang_dict)
        assert result == "en-US"

    def test_en_gb_recognized(self):
        """en-GB is recognized as English."""
        from yt_excel.youtube import _find_english_code

        lang_dict = {"en-GB": [{"ext": "vtt"}], "de": [{"ext": "vtt"}]}
        result = _find_english_code(lang_dict)
        assert result == "en-GB"

    def test_en_exact_preferred_over_variant(self):
        """'en' exact match is preferred over 'en-US'."""
        from yt_excel.youtube import _find_english_code

        lang_dict = {
            "en-US": [{"ext": "vtt"}],
            "en": [{"ext": "vtt"}],
            "en-GB": [{"ext": "vtt"}],
        }
        result = _find_english_code(lang_dict)
        assert result == "en"

    def test_en_ca_recognized(self):
        """en-CA (Canadian English) is recognized."""
        from yt_excel.youtube import _find_english_code

        lang_dict = {"en-CA": [{"ext": "vtt"}]}
        result = _find_english_code(lang_dict)
        assert result == "en-CA"

    @patch("yt_excel.cli.download_captions")
    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    @patch("yt_excel.cli.validate_api_key")
    @patch("yt_excel.translator.time.sleep")
    def test_en_us_variant_pipeline(
        self,
        mock_sleep,
        mock_api_key,
        mock_fetch_meta,
        mock_list_captions,
        mock_download,
        tmp_path,
    ):
        """Pipeline works with en-US caption variant."""
        mock_api_key.return_value = "sk-test"
        mock_fetch_meta.return_value = VideoMeta(
            video_id="enus_test12",
            title="EN-US Test",
            channel="TestChannel",
            duration="00:02:00",
        )
        mock_list_captions.return_value = CaptionInfo(
            lang_code="en-US",
            caption_type="manual",
            available_codes=["en-US"],
        )
        mock_download.return_value = SAMPLE_VTT

        args = _make_args(
            "https://www.youtube.com/watch?v=enus_test12", tmp_path
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = _mock_translation_response(5)

        with patch("yt_excel.cli.create_client", return_value=mock_client):
            with patch("yt_excel.cli.detect_font", return_value="Malgun Gothic"):
                _run_pipeline(args, config, out, 0.0)

        # Verify the download was called with en-US
        mock_download.assert_called_with("enus_test12", "en-US")


class TestDuplicateVideo:
    """Tests for duplicate video detection."""

    @patch("yt_excel.cli.download_captions")
    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    @patch("yt_excel.cli.validate_api_key")
    @patch("yt_excel.translator.time.sleep")
    def test_duplicate_video_exits_gracefully(
        self,
        mock_sleep,
        mock_api_key,
        mock_fetch_meta,
        mock_list_captions,
        mock_download,
        tmp_path,
    ):
        """Second processing of same video ID exits with info message."""
        mock_api_key.return_value = "sk-test"
        mock_fetch_meta.return_value = VideoMeta(
            video_id="dQw4w9WgXcQ",
            title="How DNA Works",
            channel="TED-Ed",
            duration="00:04:52",
        )
        mock_list_captions.return_value = CaptionInfo(
            lang_code="en", caption_type="manual", available_codes=["en"]
        )
        mock_download.return_value = SAMPLE_VTT

        args = _make_args(
            "https://www.youtube.com/watch?v=dQw4w9WgXcQ", tmp_path
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = _mock_translation_response(5)

        # First run
        with patch("yt_excel.cli.create_client", return_value=mock_client):
            with patch("yt_excel.cli.detect_font", return_value="Malgun Gothic"):
                _run_pipeline(args, config, out, 0.0)

        # Second run - should exit with code 0 (info, not error)
        with pytest.raises(SystemExit) as exc_info:
            _run_pipeline(args, config, out, 0.0)
        assert exc_info.value.code == 0


class TestDryRunMode:
    """Tests for dry-run mode."""

    @patch("yt_excel.cli.download_captions")
    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    def test_dry_run_skips_api_key_validation(
        self, mock_fetch_meta, mock_list_captions, mock_download, tmp_path
    ):
        """Dry run skips API key validation."""
        mock_fetch_meta.return_value = VideoMeta(
            video_id="dryrun12345",
            title="Dry Run Test",
            channel="TestChannel",
            duration="00:02:00",
        )
        mock_list_captions.return_value = CaptionInfo(
            lang_code="en", caption_type="manual", available_codes=["en"]
        )
        mock_download.return_value = SAMPLE_VTT

        args = _make_args(
            "https://www.youtube.com/watch?v=dryrun12345", tmp_path, dry_run=True
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        # Should not raise even without API key set
        with patch.dict(os.environ, {}, clear=False):
            _run_pipeline(args, config, out, 0.0)

        # No Excel file should be created (only if Master.xlsx didn't already exist)
        # Actually dry-run still initializes workbook for duplicate check
        # but doesn't write data sheets

    @patch("yt_excel.cli.download_captions")
    @patch("yt_excel.cli.list_captions")
    @patch("yt_excel.cli.fetch_metadata")
    def test_dry_run_no_translation_api_called(
        self, mock_fetch_meta, mock_list_captions, mock_download, tmp_path
    ):
        """Dry run does not call translation API."""
        mock_fetch_meta.return_value = VideoMeta(
            video_id="dryrun23456",
            title="Dry Run No API",
            channel="TestChannel",
            duration="00:02:00",
        )
        mock_list_captions.return_value = CaptionInfo(
            lang_code="en", caption_type="manual", available_codes=["en"]
        )
        mock_download.return_value = SAMPLE_VTT

        args = _make_args(
            "https://www.youtube.com/watch?v=dryrun23456", tmp_path, dry_run=True
        )
        config = _make_config(tmp_path)
        out = Output("quiet")

        with patch("yt_excel.cli.create_client") as mock_create_client:
            _run_pipeline(args, config, out, 0.0)
            # create_client should not be called in dry-run
            mock_create_client.assert_not_called()
