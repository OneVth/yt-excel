"""Phase 7: Excel edge case tests.

Covers:
- Sheet name 31-char truncation with special characters
- Special character titles
- Duplicate titles (same title processed twice)
- 100+ sheets in Master.xlsx
- File lock detection
- Read-only file handling
- Korean/space paths
"""

import os
from pathlib import Path

import openpyxl
import pytest

from yt_excel.excel import (
    DuplicateVideoError,
    FileLockError,
    METADATA_HEADERS,
    METADATA_SHEET,
    STUDY_LOG_SHEET,
    apply_all_styles,
    check_duplicate,
    check_file_lock,
    detect_font,
    generate_unique_sheet_name,
    initialize_workbook,
    sanitize_sheet_name,
    write_data_sheet,
    write_metadata_row,
    write_study_log_row,
    MetadataRow,
)
from yt_excel.vtt import Segment


class TestSheetNameLongTitles:
    """Tests for 31-char limit with various title lengths."""

    def test_exactly_31_chars(self):
        """Title with exactly 31 characters is unchanged."""
        title = "A" * 31
        result = sanitize_sheet_name(title)
        assert len(result) == 31
        assert result == title

    def test_32_chars_truncated_with_ellipsis(self):
        """Title with 32 characters gets truncated to 30 + ellipsis."""
        title = "A" * 32
        result = sanitize_sheet_name(title)
        assert len(result) == 31
        assert result[-1] == "\u2026"
        assert result[:-1] == "A" * 30

    def test_very_long_title_100_chars(self):
        """Very long title (100 chars) truncated to 31."""
        title = "The Incredible Journey " * 5  # >100 chars
        result = sanitize_sheet_name(title)
        assert len(result) == 31
        assert result.endswith("\u2026")

    def test_long_title_with_forbidden_chars(self):
        """Long title with forbidden chars: replaced chars may change length."""
        title = "What is 1/0? [Part 2]: A *New* Theory of Everything"
        result = sanitize_sheet_name(title)
        assert len(result) <= 31
        assert "/" not in result
        assert "?" not in result
        assert "[" not in result
        assert "]" not in result
        assert ":" not in result
        assert "*" not in result

    def test_truncation_strips_trailing_space(self):
        """If truncation ends on a space, the space is removed."""
        # Design a title where char 30 would be a space
        title = "A" * 29 + " " + "B" * 10  # Space at position 30
        result = sanitize_sheet_name(title)
        assert len(result) <= 31
        assert not result.endswith(" \u2026")


class TestSpecialCharacterTitles:
    """Tests for titles with various special characters."""

    def test_all_forbidden_chars_replaced(self):
        """All 7 forbidden chars are properly handled."""
        title = 'a/b\\c?d*e[f]g:h'
        result = sanitize_sheet_name(title)
        assert result == "a-b-cde(f)g-h"

    def test_question_mark_and_asterisk_deleted(self):
        """? and * are deleted (not replaced)."""
        title = "What? How* Why?"
        result = sanitize_sheet_name(title)
        assert result == "What How Why"

    def test_slashes_become_dashes(self):
        """/ and \\ become dashes."""
        title = "AC/DC and Back\\Slash"
        result = sanitize_sheet_name(title)
        assert result == "AC-DC and Back-Slash"

    def test_brackets_become_parentheses(self):
        """[ and ] become ( and )."""
        title = "Test [Part 1]"
        result = sanitize_sheet_name(title)
        assert result == "Test (Part 1)"

    def test_leading_trailing_quotes_stripped(self):
        """Leading and trailing single quotes are removed."""
        title = "'Hello World'"
        result = sanitize_sheet_name(title)
        assert result == "Hello World"

    def test_internal_quotes_preserved(self):
        """Internal single quotes are preserved."""
        title = "It's a wonderful life"
        result = sanitize_sheet_name(title)
        assert result == "It's a wonderful life"

    def test_consecutive_whitespace_collapsed(self):
        """Multiple spaces and tabs collapsed to single space."""
        title = "Hello   World\t\tTest"
        result = sanitize_sheet_name(title)
        assert result == "Hello World Test"

    def test_empty_after_sanitization_gets_fallback(self):
        """Title that becomes empty after sanitization gets 'Untitled'."""
        title = "??**"
        result = sanitize_sheet_name(title)
        assert result == "Untitled"

    def test_unicode_characters_preserved(self):
        """Unicode characters (Korean, emoji, etc.) are preserved."""
        title = "DNA 작동 원리"
        result = sanitize_sheet_name(title)
        assert result == "DNA 작동 원리"

    def test_design_doc_example_what_is_1_0(self):
        """Design doc example: 'What is 1/0?' -> 'What is 1-0'."""
        title = "What is 1/0?"
        result = sanitize_sheet_name(title)
        assert result == "What is 1-0"


class TestDuplicateSheetNames:
    """Tests for handling duplicate sheet names."""

    def test_first_use_no_suffix(self):
        """First use of a name gets no suffix."""
        result = generate_unique_sheet_name("Test Title", [])
        assert result == "Test Title"

    def test_second_use_gets_suffix_2(self):
        """Second use gets (2) suffix."""
        result = generate_unique_sheet_name("Test Title", ["Test Title"])
        assert result == "Test Title(2)"

    def test_third_use_gets_suffix_3(self):
        """Third use gets (3) suffix."""
        result = generate_unique_sheet_name("Test Title", ["Test Title", "Test Title(2)"])
        assert result == "Test Title(3)"

    def test_long_title_truncated_for_suffix(self):
        """Long title is further truncated to fit (2) suffix within 31 chars."""
        title = "A" * 31  # Already at max
        existing = [sanitize_sheet_name(title)]
        result = generate_unique_sheet_name(title, existing)
        assert len(result) <= 31
        assert "(2)" in result

    def test_duplicate_with_forbidden_chars(self):
        """Duplicate detection works after forbidden char replacement."""
        existing = ["What is 1-0"]
        result = generate_unique_sheet_name("What is 1/0?", existing)
        assert result == "What is 1-0(2)"

    def test_many_duplicates(self):
        """Many duplicates get incrementing suffixes."""
        existing = ["Test", "Test(2)", "Test(3)", "Test(4)", "Test(5)"]
        result = generate_unique_sheet_name("Test", existing)
        assert result == "Test(6)"


class TestManySheets:
    """Tests for workbooks with many sheets."""

    def test_100_plus_sheets(self, tmp_path):
        """Workbook with 100+ sheets remains functional."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        wb = openpyxl.load_workbook(str(master_path))

        # Create 100 data sheets
        for i in range(100):
            segments = [
                Segment(index=1, start="00:00:01.000", end="00:00:04.000",
                        english=f"Segment {i}", korean=f"세그먼트 {i}"),
            ]
            sheet_name = generate_unique_sheet_name(f"Video {i}", wb.sheetnames)
            write_data_sheet(wb, sheet_name, segments)

        wb.save(str(master_path))

        # Verify all sheets exist
        wb = openpyxl.load_workbook(str(master_path))
        # 100 data sheets + _metadata + _study_log = 102
        assert len(wb.sheetnames) == 102

    def test_duplicate_check_in_large_metadata(self, tmp_path):
        """Duplicate check works with many entries in _metadata."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        wb = openpyxl.load_workbook(str(master_path))

        # Add 50 metadata entries
        for i in range(50):
            row = MetadataRow(
                video_id=f"vid{i:07d}",
                video_title=f"Video {i}",
                video_url=f"https://youtube.com/watch?v=vid{i:07d}",
                channel_name="TestChannel",
                video_duration="00:05:00",
                sheet_name=f"Video {i}",
                processed_at="2025-01-01T00:00:00+00:00",
                total_segments=10,
                filtered_segments=2,
                translation_success=10,
                translation_failed=0,
                model_used="gpt-5-nano",
                tool_version="0.1.0",
            )
            write_metadata_row(wb, row)

        wb.save(str(master_path))

        # Check existing video is detected
        with pytest.raises(DuplicateVideoError):
            check_duplicate(str(master_path), "vid0000025")

        # Check new video passes
        check_duplicate(str(master_path), "newvideo001")


class TestFileLock:
    """Tests for file lock detection."""

    def test_writable_file_passes(self, tmp_path):
        """Writable file passes lock check."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))
        check_file_lock(str(master_path))  # Should not raise

    def test_nonexistent_file_passes(self, tmp_path):
        """Non-existent file passes lock check (nothing to lock)."""
        master_path = tmp_path / "nonexistent.xlsx"
        check_file_lock(str(master_path))  # Should not raise

    def test_permission_error_raises_file_lock_error(self, tmp_path):
        """PermissionError is caught and converted to FileLockError."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        # Simulate locked file by mocking open
        from unittest.mock import patch, mock_open

        with patch("builtins.open", side_effect=PermissionError("locked")):
            with pytest.raises(FileLockError, match="locked or read-only"):
                check_file_lock(str(master_path))


class TestReadOnlyFile:
    """Tests for read-only file handling."""

    def test_read_only_file_detected(self, tmp_path):
        """Read-only Master.xlsx triggers FileLockError."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        # Make file read-only
        os.chmod(str(master_path), 0o444)

        try:
            with pytest.raises(FileLockError):
                check_file_lock(str(master_path))
        finally:
            # Restore permissions for cleanup
            os.chmod(str(master_path), 0o666)


class TestPathsWithSpecialCharacters:
    """Tests for paths with Korean characters and spaces."""

    def test_korean_path(self, tmp_path):
        """Master.xlsx in a Korean-named directory works."""
        korean_dir = tmp_path / "한글 폴더"
        korean_dir.mkdir()
        master_path = korean_dir / "Master.xlsx"
        result = initialize_workbook(str(master_path))
        assert result.created
        assert master_path.exists()

    def test_space_in_path(self, tmp_path):
        """Master.xlsx in a path with spaces works."""
        space_dir = tmp_path / "my project folder"
        space_dir.mkdir()
        master_path = space_dir / "Master.xlsx"
        result = initialize_workbook(str(master_path))
        assert result.created
        assert master_path.exists()

    def test_korean_and_space_path(self, tmp_path):
        """Master.xlsx in path with both Korean and spaces works."""
        mixed_dir = tmp_path / "프로젝트 폴더 2025"
        mixed_dir.mkdir()
        master_path = mixed_dir / "Master.xlsx"
        result = initialize_workbook(str(master_path))
        assert result.created

        # Verify file can be opened and operated on
        wb = openpyxl.load_workbook(str(master_path))
        assert METADATA_SHEET in wb.sheetnames
        assert STUDY_LOG_SHEET in wb.sheetnames

    def test_write_and_read_in_special_path(self, tmp_path):
        """Full write cycle in a special-character path."""
        special_dir = tmp_path / "내 프로젝트"
        special_dir.mkdir()
        master_path = special_dir / "Master.xlsx"
        initialize_workbook(str(master_path))

        wb = openpyxl.load_workbook(str(master_path))
        segments = [
            Segment(index=1, start="00:00:01.000", end="00:00:04.000",
                    english="Hello world", korean="안녕하세요"),
        ]
        write_data_sheet(wb, "Test Sheet", segments)
        wb.save(str(master_path))

        # Re-read and verify
        wb2 = openpyxl.load_workbook(str(master_path))
        ws = wb2["Test Sheet"]
        assert ws.cell(row=2, column=4).value == "Hello world"
        assert ws.cell(row=2, column=5).value == "안녕하세요"


class TestStylesOnEdgeCases:
    """Tests for style application on edge-case data."""

    def test_styles_with_empty_korean(self, tmp_path):
        """Conditional formatting works with empty Korean (translation failure)."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        wb = openpyxl.load_workbook(str(master_path))
        segments = [
            Segment(index=1, start="00:00:01.000", end="00:00:04.000",
                    english="Hello", korean="안녕"),
            Segment(index=2, start="00:00:05.000", end="00:00:08.000",
                    english="World", korean=""),  # Translation failed
        ]
        write_data_sheet(wb, "Test", segments)
        apply_all_styles(wb, "Malgun Gothic", data_sheet_name="Test")
        wb.save(str(master_path))

        # Verify conditional formatting was applied
        wb2 = openpyxl.load_workbook(str(master_path))
        ws = wb2["Test"]
        assert len(ws.conditional_formatting._cf_rules) > 0

    def test_styles_with_single_segment(self, tmp_path):
        """Styles work with just one segment."""
        master_path = tmp_path / "Master.xlsx"
        initialize_workbook(str(master_path))

        wb = openpyxl.load_workbook(str(master_path))
        segments = [
            Segment(index=1, start="00:00:01.000", end="00:00:04.000",
                    english="Only segment", korean="유일한 세그먼트"),
        ]
        write_data_sheet(wb, "Single", segments)
        apply_all_styles(wb, "Malgun Gothic", data_sheet_name="Single")
        wb.save(str(master_path))

        wb2 = openpyxl.load_workbook(str(master_path))
        ws = wb2["Single"]
        assert ws.freeze_panes == "A2"
