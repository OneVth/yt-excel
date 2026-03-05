"""Tests for Excel reader/writer (Master.xlsx operations)."""

import openpyxl
import pytest

from datetime import date

from yt_excel.excel import (
    COL_WIDTH_ENGLISH,
    COL_WIDTH_INDEX,
    COL_WIDTH_KOREAN,
    COL_WIDTH_TIMESTAMP,
    DATA_HEADERS,
    DuplicateVideoError,
    FAIL_ROW_BG,
    FileLockError,
    HEADER_BG,
    HEADER_FG,
    METADATA_HEADERS,
    METADATA_SHEET,
    MetadataRow,
    NOT_STARTED_BG,
    STUDY_LOG_HEADERS,
    STUDY_LOG_SHEET,
    TAB_COLOR_METADATA,
    TAB_COLOR_STUDY_LOG,
    InitResult,
    apply_all_styles,
    apply_data_sheet_style,
    apply_metadata_style,
    apply_study_log_style,
    check_duplicate,
    check_file_lock,
    detect_font,
    generate_unique_sheet_name,
    initialize_workbook,
    sanitize_sheet_name,
    write_data_sheet,
    write_metadata_row,
    write_study_log_row,
)
from yt_excel.vtt import Segment


class TestInitializeWorkbook:
    """Tests for initialize_workbook — creation and integrity checks."""

    def test_creates_new_file_when_not_exists(self, tmp_path):
        """New file is created with _metadata and _study_log sheets."""
        path = tmp_path / "Master.xlsx"
        result = initialize_workbook(path)

        assert result.created is True
        assert result.metadata_recovered is False
        assert result.study_log_recovered is False
        assert path.exists()

        wb = openpyxl.load_workbook(str(path))
        assert METADATA_SHEET in wb.sheetnames
        assert STUDY_LOG_SHEET in wb.sheetnames
        # Should not have the default "Sheet"
        assert "Sheet" not in wb.sheetnames

    def test_new_file_has_metadata_headers(self, tmp_path):
        """Newly created _metadata sheet has all 13 column headers."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)

        wb = openpyxl.load_workbook(str(path))
        ws = wb[METADATA_SHEET]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 14)]
        assert headers == METADATA_HEADERS

    def test_new_file_has_study_log_headers(self, tmp_path):
        """Newly created _study_log sheet has all 8 column headers."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)

        wb = openpyxl.load_workbook(str(path))
        ws = wb[STUDY_LOG_SHEET]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 9)]
        assert headers == STUDY_LOG_HEADERS

    def test_existing_file_both_sheets_intact(self, tmp_path):
        """Existing file with both sheets passes integrity check unchanged."""
        path = tmp_path / "Master.xlsx"
        # Pre-create a valid file
        initialize_workbook(path)

        # Run again — should detect both sheets and do nothing
        result = initialize_workbook(path)
        assert result.created is False
        assert result.metadata_recovered is False
        assert result.study_log_recovered is False

    def test_recovers_missing_study_log(self, tmp_path):
        """Missing _study_log is recreated while preserving _metadata."""
        path = tmp_path / "Master.xlsx"
        # Create file, then remove _study_log
        wb = openpyxl.Workbook()
        default = wb.active
        if default is not None:
            wb.remove(default)
        ws_meta = wb.create_sheet(METADATA_SHEET)
        ws_meta.cell(row=1, column=1, value="video_id")
        ws_meta.cell(row=2, column=1, value="test123abcd")
        wb.save(str(path))

        result = initialize_workbook(path)
        assert result.created is False
        assert result.metadata_recovered is False
        assert result.study_log_recovered is True

        wb = openpyxl.load_workbook(str(path))
        assert STUDY_LOG_SHEET in wb.sheetnames
        # Existing _metadata data should be preserved
        ws = wb[METADATA_SHEET]
        assert ws.cell(row=2, column=1).value == "test123abcd"

    def test_recovers_missing_metadata(self, tmp_path):
        """Missing _metadata is recreated while preserving other sheets."""
        path = tmp_path / "Master.xlsx"
        wb = openpyxl.Workbook()
        default = wb.active
        if default is not None:
            wb.remove(default)
        ws_log = wb.create_sheet(STUDY_LOG_SHEET)
        ws_log.cell(row=1, column=1, value="No")
        ws_log.cell(row=2, column=1, value=1)
        # Also add a data sheet
        ws_data = wb.create_sheet("How DNA Works")
        ws_data.cell(row=1, column=1, value="Index")
        ws_data.cell(row=2, column=1, value=1)
        wb.save(str(path))

        result = initialize_workbook(path)
        assert result.metadata_recovered is True
        assert result.study_log_recovered is False

        wb = openpyxl.load_workbook(str(path))
        assert METADATA_SHEET in wb.sheetnames
        # Existing _study_log and data sheet preserved
        assert wb[STUDY_LOG_SHEET].cell(row=2, column=1).value == 1
        assert wb["How DNA Works"].cell(row=2, column=1).value == 1

    def test_recovers_both_missing_sheets(self, tmp_path):
        """Both _metadata and _study_log recreated when both missing."""
        path = tmp_path / "Master.xlsx"
        wb = openpyxl.Workbook()
        # Just has a data sheet
        ws = wb.active
        ws.title = "Some Data Sheet"
        ws.cell(row=1, column=1, value="Index")
        wb.save(str(path))

        result = initialize_workbook(path)
        assert result.metadata_recovered is True
        assert result.study_log_recovered is True

        wb = openpyxl.load_workbook(str(path))
        assert METADATA_SHEET in wb.sheetnames
        assert STUDY_LOG_SHEET in wb.sheetnames
        # Original data sheet still there
        assert "Some Data Sheet" in wb.sheetnames


class TestCheckFileLock:
    """Tests for check_file_lock — best-effort write permission check."""

    def test_writable_file_passes(self, tmp_path):
        """Normal writable file does not raise."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        # Should not raise
        check_file_lock(path)

    def test_nonexistent_file_passes(self, tmp_path):
        """Non-existent file does not raise (nothing to lock)."""
        path = tmp_path / "Master.xlsx"
        check_file_lock(path)

    def test_locked_file_raises(self, tmp_path, monkeypatch):
        """PermissionError when opening file raises FileLockError."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)

        # Mock builtins.open to raise PermissionError (simulates Excel lock)
        import builtins
        real_open = builtins.open

        def mock_open(p, *args, **kwargs):
            if str(p) == str(path) and "r+b" in args:
                raise PermissionError("File is locked")
            return real_open(p, *args, **kwargs)

        monkeypatch.setattr(builtins, "open", mock_open)

        with pytest.raises(FileLockError, match="locked or read-only"):
            check_file_lock(path)


class TestCheckDuplicate:
    """Tests for check_duplicate — video_id lookup in _metadata."""

    def test_new_video_passes(self, tmp_path):
        """New video_id does not raise."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        # Should not raise
        check_duplicate(path, "abcde123456")

    def test_nonexistent_file_passes(self, tmp_path):
        """Non-existent file does not raise."""
        path = tmp_path / "Master.xlsx"
        check_duplicate(path, "abcde123456")

    def test_duplicate_video_raises(self, tmp_path):
        """Existing video_id raises DuplicateVideoError with details."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)

        # Add a row to _metadata
        wb = openpyxl.load_workbook(str(path))
        ws = wb[METADATA_SHEET]
        ws.cell(row=2, column=1, value="dQw4w9WgXcQ")
        ws.cell(row=2, column=6, value="How DNA Works")
        ws.cell(row=2, column=7, value="2026-03-01T12:00:00")
        wb.save(str(path))

        with pytest.raises(DuplicateVideoError) as exc_info:
            check_duplicate(path, "dQw4w9WgXcQ")

        assert exc_info.value.video_id == "dQw4w9WgXcQ"
        assert exc_info.value.sheet_name == "How DNA Works"
        assert exc_info.value.processed_at == "2026-03-01T12:00:00"

    def test_different_video_passes(self, tmp_path):
        """Different video_id does not raise when other videos exist."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)

        wb = openpyxl.load_workbook(str(path))
        ws = wb[METADATA_SHEET]
        ws.cell(row=2, column=1, value="existingID01")
        wb.save(str(path))

        # Different ID should pass
        check_duplicate(path, "newVideoID01")

    def test_missing_metadata_sheet_passes(self, tmp_path):
        """File without _metadata sheet does not raise."""
        path = tmp_path / "Master.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(path))

        check_duplicate(path, "anyVideoID01")


class TestSanitizeSheetName:
    """Tests for sanitize_sheet_name — character replacement and truncation."""

    def test_normal_title_unchanged(self):
        """Short title without special chars passes through unchanged."""
        assert sanitize_sheet_name("How DNA Works") == "How DNA Works"

    def test_slash_replaced(self):
        """Forward slash is replaced with dash."""
        assert sanitize_sheet_name("What is 1/0?") == "What is 1-0"

    def test_backslash_replaced(self):
        """Backslash is replaced with dash."""
        assert sanitize_sheet_name("A\\B") == "A-B"

    def test_question_mark_removed(self):
        """Question mark is deleted."""
        assert sanitize_sheet_name("Why?") == "Why"

    def test_asterisk_removed(self):
        """Asterisk is deleted."""
        assert sanitize_sheet_name("A *New* Theory") == "A New Theory"

    def test_brackets_replaced(self):
        """Square brackets replaced with parentheses."""
        assert sanitize_sheet_name("[Part 2]") == "(Part 2)"

    def test_colon_replaced(self):
        """Colon replaced with dash."""
        assert sanitize_sheet_name("Topic: Intro") == "Topic- Intro"

    def test_combined_special_chars(self):
        """Multiple special chars in one title."""
        result = sanitize_sheet_name("What is 1/0? [Part 2]: A *New* Theory")
        assert result == "What is 1-0 (Part 2)- A New Th\u2026"
        assert len(result) <= 31

    def test_truncation_at_31_chars(self):
        """Titles longer than 31 chars are truncated with ellipsis."""
        title = "The Incredible Journey of a Red Blood Cell"
        result = sanitize_sheet_name(title)
        assert len(result) == 31
        assert result.endswith("\u2026")
        assert result == "The Incredible Journey of a Re\u2026"

    def test_trailing_space_stripped_after_truncation(self):
        """Trailing space before ellipsis is removed."""
        # 30th char would be a space: "X" * 29 + " " + "more"
        title = "A" * 29 + " " + "more text here"
        result = sanitize_sheet_name(title)
        assert len(result) <= 31
        assert not result[-2:-1].isspace()

    def test_leading_trailing_quotes_stripped(self):
        """Leading and trailing single quotes are removed."""
        assert sanitize_sheet_name("'Hello World'") == "Hello World"

    def test_empty_title_fallback(self):
        """Empty or all-forbidden title gets fallback name."""
        assert sanitize_sheet_name("???") == "Untitled"
        assert sanitize_sheet_name("") == "Untitled"

    def test_consecutive_whitespace_collapsed(self):
        """Multiple spaces/tabs are collapsed to single space."""
        assert sanitize_sheet_name("A   B\tC") == "A B C"


class TestGenerateUniqueSheetName:
    """Tests for generate_unique_sheet_name — deduplication with suffix."""

    def test_unique_name_no_suffix(self):
        """Unique name returns without suffix."""
        result = generate_unique_sheet_name("How DNA Works", [])
        assert result == "How DNA Works"

    def test_duplicate_gets_suffix_2(self):
        """First duplicate gets (2) suffix."""
        existing = ["How DNA Works"]
        result = generate_unique_sheet_name("How DNA Works", existing)
        assert result == "How DNA Works(2)"

    def test_multiple_duplicates(self):
        """Multiple duplicates increment suffix correctly."""
        existing = ["How DNA Works", "How DNA Works(2)"]
        result = generate_unique_sheet_name("How DNA Works", existing)
        assert result == "How DNA Works(3)"

    def test_long_title_with_suffix_stays_within_31(self):
        """Long title is further truncated to accommodate suffix."""
        title = "The Incredible Journey of a Red Blood Cell"
        # First one gets normal truncation
        first = sanitize_sheet_name(title)
        assert len(first) == 31

        # Second one needs suffix, base must be shortened
        result = generate_unique_sheet_name(title, [first])
        assert len(result) <= 31
        assert result.endswith("(2)")

    def test_suffix_with_exact_31_char_name(self):
        """Name exactly at 31 chars gets truncated for suffix."""
        title = "A" * 31
        existing = [sanitize_sheet_name(title)]
        result = generate_unique_sheet_name(title, existing)
        assert len(result) <= 31
        assert "(2)" in result


def _make_segments(count=3):
    """Helper to create test segments."""
    return [
        Segment(
            index=i + 1,
            start=f"00:00:{i:02d}.000",
            end=f"00:00:{i + 1:02d}.000",
            english=f"Hello world {i + 1}",
            korean=f"안녕 세계 {i + 1}",
        )
        for i in range(count)
    ]


class TestWriteDataSheet:
    """Tests for write_data_sheet — segment data writing."""

    def test_writes_headers(self, tmp_path):
        """Data sheet has correct column headers."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        segments = _make_segments(1)
        write_data_sheet(wb, "Test Sheet", segments)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Test Sheet"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 6)]
        assert headers == DATA_HEADERS

    def test_writes_segment_data(self, tmp_path):
        """Each segment is written to the correct row and column."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        segments = _make_segments(3)
        write_data_sheet(wb, "Test Sheet", segments)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Test Sheet"]

        # Check row 2 (segment 1)
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "00:00:00.000"
        assert ws.cell(row=2, column=3).value == "00:00:01.000"
        assert ws.cell(row=2, column=4).value == "Hello world 1"
        assert ws.cell(row=2, column=5).value == "안녕 세계 1"

        # Check row 4 (segment 3)
        assert ws.cell(row=4, column=1).value == 3
        assert ws.cell(row=4, column=4).value == "Hello world 3"

    def test_timestamp_stored_as_text(self, tmp_path):
        """Timestamp cells use text format (@) to prevent auto-conversion."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        segments = _make_segments(1)
        write_data_sheet(wb, "Test Sheet", segments)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Test Sheet"]

        # Start and End cells should be text format
        assert ws.cell(row=2, column=2).number_format == "@"
        assert ws.cell(row=2, column=3).number_format == "@"
        # Values should be strings
        assert isinstance(ws.cell(row=2, column=2).value, str)
        assert isinstance(ws.cell(row=2, column=3).value, str)

    def test_empty_korean_for_failed_translation(self, tmp_path):
        """Segments with empty korean field are stored correctly."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        segments = [Segment(
            index=1,
            start="00:00:01.000",
            end="00:00:02.000",
            english="Some text",
            korean="",
        )]
        write_data_sheet(wb, "Test Sheet", segments)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Test Sheet"]
        # openpyxl stores empty strings as None
        assert ws.cell(row=2, column=5).value is None or ws.cell(row=2, column=5).value == ""


def _make_metadata_row(**overrides) -> MetadataRow:
    """Helper to create a MetadataRow with defaults."""
    defaults = {
        "video_id": "dQw4w9WgXcQ",
        "video_title": "How DNA Works",
        "video_url": "https://www.youtube.com/watch?v=dQw4w9WgXcQ",
        "channel_name": "TED-Ed",
        "video_duration": "00:04:52",
        "sheet_name": "How DNA Works",
        "processed_at": "2026-03-05T12:00:00",
        "total_segments": 131,
        "filtered_segments": 11,
        "translation_success": 131,
        "translation_failed": 0,
        "model_used": "gpt-5-nano",
        "tool_version": "0.1.0",
    }
    defaults.update(overrides)
    return MetadataRow(**defaults)


class TestWriteMetadataRow:
    """Tests for write_metadata_row — appending to _metadata sheet."""

    def test_writes_all_13_fields(self, tmp_path):
        """All 13 metadata fields are written correctly."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        row_data = _make_metadata_row()
        write_metadata_row(wb, row_data)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb[METADATA_SHEET]

        assert ws.cell(row=2, column=1).value == "dQw4w9WgXcQ"
        assert ws.cell(row=2, column=2).value == "How DNA Works"
        assert ws.cell(row=2, column=3).value == "https://www.youtube.com/watch?v=dQw4w9WgXcQ"
        assert ws.cell(row=2, column=4).value == "TED-Ed"
        assert ws.cell(row=2, column=5).value == "00:04:52"
        assert ws.cell(row=2, column=6).value == "How DNA Works"
        assert ws.cell(row=2, column=7).value == "2026-03-05T12:00:00"
        assert ws.cell(row=2, column=8).value == 131
        assert ws.cell(row=2, column=9).value == 11
        assert ws.cell(row=2, column=10).value == 131
        assert ws.cell(row=2, column=11).value == 0
        assert ws.cell(row=2, column=12).value == "gpt-5-nano"
        assert ws.cell(row=2, column=13).value == "0.1.0"

    def test_appends_to_next_row(self, tmp_path):
        """Multiple metadata rows are appended sequentially."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        write_metadata_row(wb, _make_metadata_row(video_id="first_vid_01"))
        write_metadata_row(wb, _make_metadata_row(video_id="second_vid02"))
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb[METADATA_SHEET]
        assert ws.cell(row=2, column=1).value == "first_vid_01"
        assert ws.cell(row=3, column=1).value == "second_vid02"


class TestWriteStudyLogRow:
    """Tests for write_study_log_row — appending to _study_log sheet."""

    def test_writes_auto_fields_and_defaults(self, tmp_path):
        """Writes No, Study Date, Title, Duration, Segments, Status, Review Count."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        write_study_log_row(wb, "How DNA Works", "00:04:52", 131)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb[STUDY_LOG_SHEET]

        assert ws.cell(row=2, column=1).value == 1  # No
        assert ws.cell(row=2, column=2).value == date.today().isoformat()  # Study Date
        assert ws.cell(row=2, column=3).value == "How DNA Works"  # Video Title
        assert ws.cell(row=2, column=4).value == "04:52"  # Duration (MM:SS)
        assert ws.cell(row=2, column=5).value == 131  # Segments
        assert ws.cell(row=2, column=6).value == "Not Started"  # Status
        assert ws.cell(row=2, column=7).value == 0  # Review Count
        # Notes should be empty
        assert ws.cell(row=2, column=8).value is None

    def test_auto_increments_no(self, tmp_path):
        """No field auto-increments for each new row."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        write_study_log_row(wb, "Video 1", "00:05:00", 100)
        write_study_log_row(wb, "Video 2", "00:10:00", 200)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb[STUDY_LOG_SHEET]
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=3, column=1).value == 2

    def test_duration_format_conversion(self, tmp_path):
        """HH:MM:SS duration is converted to MM:SS."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        write_study_log_row(wb, "Long Video", "01:30:45", 500)
        wb.save(str(path))

        wb = openpyxl.load_workbook(str(path))
        ws = wb[STUDY_LOG_SHEET]
        # 1 hour 30 minutes = 90 minutes
        assert ws.cell(row=2, column=4).value == "90:45"


class TestDetectFont:
    """Tests for detect_font — font auto-detection."""

    def test_explicit_font_returned(self):
        """Non-auto setting returns the specified font."""
        assert detect_font("Arial") == "Arial"

    def test_auto_returns_string(self):
        """Auto mode returns a non-empty string."""
        result = detect_font("auto")
        assert isinstance(result, str)
        assert len(result) > 0


class TestStyleEngine:
    """Tests for style application — headers, formatting, freeze, tab colors."""

    def _make_styled_workbook(self, tmp_path, segment_count=3):
        """Helper: create workbook with data, metadata, study log, then apply styles."""
        path = tmp_path / "Master.xlsx"
        initialize_workbook(path)
        wb = openpyxl.load_workbook(str(path))

        segments = _make_segments(segment_count)
        write_data_sheet(wb, "Test Sheet", segments)
        write_metadata_row(wb, _make_metadata_row())
        write_study_log_row(wb, "How DNA Works", "00:04:52", segment_count)

        font_name = "Malgun Gothic"
        apply_all_styles(wb, font_name, data_sheet_name="Test Sheet")
        wb.save(str(path))

        return openpyxl.load_workbook(str(path))

    def test_data_sheet_header_bg_color(self, tmp_path):
        """Data sheet header has correct background color."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb == "00" + HEADER_BG

    def test_data_sheet_header_font_bold(self, tmp_path):
        """Data sheet header font is bold."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True

    def test_data_sheet_header_font_color(self, tmp_path):
        """Data sheet header font has correct color."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        cell = ws.cell(row=1, column=1)
        assert cell.font.color.rgb == "00" + HEADER_FG

    def test_data_sheet_freeze_panes(self, tmp_path):
        """Data sheet has freeze panes at A2."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        assert ws.freeze_panes == "A2"

    def test_data_sheet_column_widths(self, tmp_path):
        """Data sheet has correct column widths."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        assert ws.column_dimensions["A"].width == COL_WIDTH_INDEX
        assert ws.column_dimensions["B"].width == COL_WIDTH_TIMESTAMP
        assert ws.column_dimensions["C"].width == COL_WIDTH_TIMESTAMP
        assert ws.column_dimensions["D"].width == COL_WIDTH_ENGLISH
        assert ws.column_dimensions["E"].width == COL_WIDTH_KOREAN

    def test_data_sheet_conditional_formatting_exists(self, tmp_path):
        """Data sheet has conditional formatting for failed translations."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        assert len(ws.conditional_formatting) > 0

    def test_metadata_freeze_panes(self, tmp_path):
        """_metadata sheet has freeze panes at A2."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[METADATA_SHEET]
        assert ws.freeze_panes == "A2"

    def test_metadata_auto_filter(self, tmp_path):
        """_metadata sheet has auto filter enabled."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[METADATA_SHEET]
        assert ws.auto_filter.ref is not None

    def test_metadata_tab_color(self, tmp_path):
        """_metadata sheet tab color is grey."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[METADATA_SHEET]
        assert ws.sheet_properties.tabColor.rgb == "00" + TAB_COLOR_METADATA

    def test_study_log_freeze_panes(self, tmp_path):
        """_study_log sheet has freeze panes at A2."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[STUDY_LOG_SHEET]
        assert ws.freeze_panes == "A2"

    def test_study_log_auto_filter(self, tmp_path):
        """_study_log sheet has auto filter enabled."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[STUDY_LOG_SHEET]
        assert ws.auto_filter.ref is not None

    def test_study_log_tab_color(self, tmp_path):
        """_study_log sheet tab color is blue."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[STUDY_LOG_SHEET]
        assert ws.sheet_properties.tabColor.rgb == "00" + TAB_COLOR_STUDY_LOG

    def test_study_log_conditional_formatting(self, tmp_path):
        """_study_log sheet has conditional formatting for Not Started."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[STUDY_LOG_SHEET]
        assert len(ws.conditional_formatting) > 0

    def test_data_body_wrap_text(self, tmp_path):
        """Data cells have wrap text enabled."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        cell = ws.cell(row=2, column=4)  # English column
        assert cell.alignment.wrap_text is True

    def test_data_body_vertical_top(self, tmp_path):
        """Data cells have vertical alignment set to top."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb["Test Sheet"]
        cell = ws.cell(row=2, column=4)  # English column
        assert cell.alignment.vertical == "top"

    def test_metadata_header_style(self, tmp_path):
        """_metadata header has correct background and bold font."""
        wb = self._make_styled_workbook(tmp_path)
        ws = wb[METADATA_SHEET]
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb == "00" + HEADER_BG
        assert cell.font.bold is True
