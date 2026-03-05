"""Excel reader/writer for Master.xlsx — initialization, integrity, and data operations."""

import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class FileLockError(Exception):
    """Master.xlsx is locked or read-only."""


class DuplicateVideoError(Exception):
    """Video has already been processed."""

    def __init__(self, video_id: str, sheet_name: str, processed_at: str) -> None:
        self.video_id = video_id
        self.sheet_name = sheet_name
        self.processed_at = processed_at
        super().__init__(
            f"This video has already been processed.\n"
            f"Sheet: {sheet_name}\n"
            f"Processed at: {processed_at}"
        )


# --- Sheet Names ---
METADATA_SHEET = "_metadata"
STUDY_LOG_SHEET = "_study_log"

# --- _metadata Column Headers (13 fields) ---
METADATA_HEADERS = [
    "video_id",
    "video_title",
    "video_url",
    "channel_name",
    "video_duration",
    "sheet_name",
    "processed_at",
    "total_segments",
    "filtered_segments",
    "translation_success",
    "translation_failed",
    "model_used",
    "tool_version",
]

# --- _study_log Column Headers (8 fields) ---
STUDY_LOG_HEADERS = [
    "No",
    "Study Date",
    "Video Title",
    "Duration",
    "Segments",
    "Status",
    "Review Count",
    "Notes",
]

# --- Data Sheet Column Headers ---
DATA_HEADERS = ["Index", "Start", "End", "English", "Korean"]

# --- Style Constants (from CLAUDE.md) ---

# Colors
HEADER_BG = "F2F2F2"
HEADER_FG = "333333"
BORDER_COLOR = "E0E0E0"
HEADER_BORDER_COLOR = "BFBFBF"
FAIL_ROW_BG = "FFF2F2"
NOT_STARTED_BG = "FFF8E1"

# Font sizes
FONT_SIZE_HEADER = 11
FONT_SIZE_BODY = 10

# Column widths — data sheet
COL_WIDTH_INDEX = 7
COL_WIDTH_TIMESTAMP = 13
COL_WIDTH_ENGLISH = 50
COL_WIDTH_KOREAN = 45

# Column widths — _metadata sheet
_METADATA_COL_WIDTHS = [14, 35, 40, 20, 12, 30, 20, 10, 10, 10, 10, 15, 12]

# Column widths — _study_log sheet
_STUDY_LOG_COL_WIDTHS = [6, 14, 35, 10, 10, 14, 14, 40]

# Tab colors
TAB_COLOR_METADATA = "808080"
TAB_COLOR_STUDY_LOG = "4472C4"


@dataclass
class InitResult:
    """Result of Master.xlsx initialization.

    Attributes:
        created: True if the file was newly created.
        metadata_recovered: True if _metadata was missing and recreated.
        study_log_recovered: True if _study_log was missing and recreated.
    """

    created: bool = False
    metadata_recovered: bool = False
    study_log_recovered: bool = False


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write header row to a worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)


def initialize_workbook(master_path: str | Path) -> InitResult:
    """Initialize Master.xlsx, creating it if missing or recovering missing sheets.

    If the file does not exist, creates a new workbook with _metadata and
    _study_log sheets (headers only). If the file exists, verifies that both
    required sheets are present and recreates any missing ones.

    Existing data sheets and their contents are never modified.

    Args:
        master_path: Path to the Master.xlsx file.

    Returns:
        InitResult describing what actions were taken.
    """
    path = Path(master_path)
    result = InitResult()

    if not path.exists():
        # Create new workbook
        wb = Workbook()

        # Remove default "Sheet" created by openpyxl
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        # Create _metadata sheet with headers
        ws_meta = wb.create_sheet(METADATA_SHEET)
        _write_header_row(ws_meta, METADATA_HEADERS)

        # Create _study_log sheet with headers
        ws_log = wb.create_sheet(STUDY_LOG_SHEET)
        _write_header_row(ws_log, STUDY_LOG_HEADERS)

        wb.save(str(path))
        result.created = True
        return result

    # File exists — verify integrity
    wb = openpyxl.load_workbook(str(path))
    modified = False

    if METADATA_SHEET not in wb.sheetnames:
        ws_meta = wb.create_sheet(METADATA_SHEET)
        _write_header_row(ws_meta, METADATA_HEADERS)
        result.metadata_recovered = True
        modified = True

    if STUDY_LOG_SHEET not in wb.sheetnames:
        ws_log = wb.create_sheet(STUDY_LOG_SHEET)
        _write_header_row(ws_log, STUDY_LOG_HEADERS)
        result.study_log_recovered = True
        modified = True

    if modified:
        wb.save(str(path))

    return result


def check_file_lock(master_path: str | Path) -> None:
    """Verify that Master.xlsx is writable (best-effort lock check).

    Attempts to open the file in append mode to detect if another process
    (e.g. Excel) holds a lock. This is best-effort — a lock acquired after
    this check cannot be prevented.

    Args:
        master_path: Path to the Master.xlsx file.

    Raises:
        FileLockError: If the file cannot be opened for writing.
    """
    path = Path(master_path)
    if not path.exists():
        return

    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise FileLockError(
            f"Master.xlsx is locked or read-only.\n"
            f"Please close the file in Excel and retry."
        )


def check_duplicate(master_path: str | Path, video_id: str) -> None:
    """Check if a video has already been processed in _metadata.

    Scans the video_id column (column 1) of the _metadata sheet.

    Args:
        master_path: Path to the Master.xlsx file.
        video_id: YouTube video ID to check.

    Raises:
        DuplicateVideoError: If the video_id already exists in _metadata.
    """
    path = Path(master_path)
    if not path.exists():
        return

    wb = openpyxl.load_workbook(str(path), read_only=True)

    if METADATA_SHEET not in wb.sheetnames:
        wb.close()
        return

    ws = wb[METADATA_SHEET]
    # Find video_id and corresponding sheet_name and processed_at columns
    # Headers are in row 1; video_id=col1, sheet_name=col6, processed_at=col7
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_video_id = row[0].value
        if cell_video_id == video_id:
            sheet_name = row[5].value if len(row) > 5 else ""
            processed_at = row[6].value if len(row) > 6 else ""
            wb.close()
            raise DuplicateVideoError(
                video_id=video_id,
                sheet_name=str(sheet_name or ""),
                processed_at=str(processed_at or ""),
            )

    wb.close()


# --- Sheet Naming ---

# Excel sheet name maximum length
_MAX_SHEET_NAME_LEN = 31

# Forbidden characters in Excel sheet names and their replacements
_CHAR_REPLACEMENTS = {
    "/": "-",
    "\\": "-",
    "?": "",
    "*": "",
    "[": "(",
    "]": ")",
    ":": "-",
}

# Regex matching any forbidden character
_FORBIDDEN_CHARS_RE = re.compile(r"[/\\?*\[\]:]")


def sanitize_sheet_name(title: str) -> str:
    """Convert a video title into a valid Excel sheet name.

    Applies the naming rules from design doc section 8.3:
    1. Replace forbidden characters (/ \\ ? * [ ] :)
    2. Collapse consecutive whitespace to single space
    3. Trim leading/trailing whitespace
    4. Strip leading/trailing single quotes
    5. Truncate to 31 characters (30 + ellipsis if needed)

    Args:
        title: Original video title.

    Returns:
        Sanitized sheet name, max 31 characters.
    """
    # 1. Replace forbidden characters
    result = _FORBIDDEN_CHARS_RE.sub(lambda m: _CHAR_REPLACEMENTS[m.group()], title)

    # 2. Collapse whitespace
    result = re.sub(r"\s+", " ", result)

    # 3. Trim
    result = result.strip()

    # 4. Strip leading/trailing single quotes
    result = result.strip("'")

    # 5. Truncate if over 31 chars
    if len(result) > _MAX_SHEET_NAME_LEN:
        result = result[:_MAX_SHEET_NAME_LEN - 1].rstrip() + "\u2026"

    # Safety: if empty after sanitization, use fallback
    if not result:
        result = "Untitled"

    return result


def generate_unique_sheet_name(
    title: str,
    existing_names: list[str],
) -> str:
    """Generate a unique sheet name, appending (2), (3), etc. if needed.

    If the sanitized name already exists among existing_names, appends
    a numeric suffix. The suffix is included within the 31-char limit.

    Args:
        title: Original video title.
        existing_names: List of existing sheet names in the workbook.

    Returns:
        Unique sheet name, max 31 characters.
    """
    base_name = sanitize_sheet_name(title)

    if base_name not in existing_names:
        return base_name

    # Try incrementing suffix
    counter = 2
    while True:
        suffix = f"({counter})"
        max_base_len = _MAX_SHEET_NAME_LEN - len(suffix)

        if len(base_name) > max_base_len:
            # Re-truncate base to fit suffix
            truncated = base_name[:max_base_len - 1].rstrip() + "\u2026"
        else:
            truncated = base_name

        candidate = f"{truncated}{suffix}"

        if candidate not in existing_names:
            return candidate

        counter += 1


# --- Data Sheet Writer ---


def write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    segments: list,
) -> Worksheet:
    """Create a data sheet and write segment rows.

    Creates the sheet with headers (Index, Start, End, English, Korean)
    and writes one row per segment. Timestamps are stored as text to
    prevent Excel from auto-converting to time format.

    Args:
        wb: The workbook to add the sheet to.
        sheet_name: Name for the new sheet (must be pre-sanitized).
        segments: List of Segment objects with index, start, end, english, korean.

    Returns:
        The created worksheet.
    """
    ws = wb.create_sheet(sheet_name)

    # Write headers
    _write_header_row(ws, DATA_HEADERS)

    # Write data rows
    for seg in segments:
        row_num = seg.index + 1  # +1 for header row

        ws.cell(row=row_num, column=1, value=seg.index)

        # Timestamps stored as text — set number format to prevent conversion
        start_cell = ws.cell(row=row_num, column=2, value=seg.start)
        start_cell.number_format = "@"

        end_cell = ws.cell(row=row_num, column=3, value=seg.end)
        end_cell.number_format = "@"

        ws.cell(row=row_num, column=4, value=seg.english)
        ws.cell(row=row_num, column=5, value=seg.korean)

    return ws


# --- Metadata Writer ---


@dataclass
class MetadataRow:
    """Data for a single _metadata row (13 fields).

    Attributes correspond to METADATA_HEADERS columns in order.
    """

    video_id: str
    video_title: str
    video_url: str
    channel_name: str
    video_duration: str
    sheet_name: str
    processed_at: str  # ISO 8601
    total_segments: int
    filtered_segments: int
    translation_success: int
    translation_failed: int
    model_used: str
    tool_version: str


def write_metadata_row(wb: Workbook, row_data: MetadataRow) -> None:
    """Append a row to the _metadata sheet.

    Args:
        wb: Workbook containing _metadata sheet.
        row_data: MetadataRow with all 13 fields.
    """
    ws = wb[METADATA_SHEET]
    next_row = ws.max_row + 1

    values = [
        row_data.video_id,
        row_data.video_title,
        row_data.video_url,
        row_data.channel_name,
        row_data.video_duration,
        row_data.sheet_name,
        row_data.processed_at,
        row_data.total_segments,
        row_data.filtered_segments,
        row_data.translation_success,
        row_data.translation_failed,
        row_data.model_used,
        row_data.tool_version,
    ]

    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)


# --- Study Log Writer ---


def _format_duration_mmss(duration_hhmmss: str) -> str:
    """Convert HH:MM:SS to MM:SS format for study log."""
    parts = duration_hhmmss.split(":")
    if len(parts) == 3:
        hours = int(parts[0])
        minutes = int(parts[1]) + hours * 60
        seconds = parts[2]
        return f"{minutes:02d}:{seconds}"
    return duration_hhmmss


def write_study_log_row(
    wb: Workbook,
    video_title: str,
    video_duration: str,
    total_segments: int,
) -> None:
    """Append a row to the _study_log sheet.

    CLI only writes the first 5 auto-generated fields (No, Study Date,
    Video Title, Duration, Segments) plus default values for Status and
    Review Count. Notes is left empty. CLI never reads or modifies
    existing rows.

    Args:
        wb: Workbook containing _study_log sheet.
        video_title: Video title string.
        video_duration: Duration in HH:MM:SS format.
        total_segments: Number of segments after filtering.
    """
    ws = wb[STUDY_LOG_SHEET]
    next_row = ws.max_row + 1

    # Auto-increment No (find last No value)
    last_no = 0
    if next_row > 2:
        prev_no = ws.cell(row=next_row - 1, column=1).value
        if isinstance(prev_no, (int, float)):
            last_no = int(prev_no)
    new_no = last_no + 1

    ws.cell(row=next_row, column=1, value=new_no)
    ws.cell(row=next_row, column=2, value=date.today().isoformat())
    ws.cell(row=next_row, column=3, value=video_title)
    ws.cell(row=next_row, column=4, value=_format_duration_mmss(video_duration))
    ws.cell(row=next_row, column=5, value=total_segments)
    ws.cell(row=next_row, column=6, value="Not Started")
    ws.cell(row=next_row, column=7, value=0)
    # Column 8 (Notes) left empty


# --- Font Detection ---


def detect_font(font_setting: str = "auto") -> str:
    """Detect the best available Korean font.

    If font_setting is "auto", checks for Noto Sans KR first,
    then falls back to Malgun Gothic (Windows default).
    Otherwise, returns the specified font name directly.

    Args:
        font_setting: Font name or "auto" for auto-detection.

    Returns:
        Font family name to use in Excel.
    """
    if font_setting != "auto":
        return font_setting

    # Check for Noto Sans KR in Windows fonts directory
    import os

    fonts_dir = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts")
    try:
        font_files = os.listdir(fonts_dir)
        for f in font_files:
            if f.lower().startswith("notosanskr"):
                return "Noto Sans KR"
    except OSError:
        pass

    return "Malgun Gothic"


# --- Style Engine ---


def _make_header_font(font_name: str) -> Font:
    """Create header font style."""
    return Font(name=font_name, size=FONT_SIZE_HEADER, bold=True, color=HEADER_FG)


def _make_body_font(font_name: str) -> Font:
    """Create body text font style."""
    return Font(name=font_name, size=FONT_SIZE_BODY)


def _make_header_fill() -> PatternFill:
    """Create header background fill."""
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")


def _make_header_border() -> Border:
    """Create header bottom border."""
    return Border(bottom=Side(style="thin", color=HEADER_BORDER_COLOR))


def _make_data_border() -> Border:
    """Create data row horizontal border (bottom only)."""
    return Border(bottom=Side(style="thin", color=BORDER_COLOR))


def apply_header_style(ws: Worksheet, font_name: str, col_count: int) -> None:
    """Apply header style to row 1 of a worksheet.

    Args:
        ws: Target worksheet.
        font_name: Font family name.
        col_count: Number of columns to style.
    """
    header_font = _make_header_font(font_name)
    header_fill = _make_header_fill()
    header_border = _make_header_border()

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="top")


def apply_data_style(
    ws: Worksheet,
    font_name: str,
    col_count: int,
    row_count: int,
) -> None:
    """Apply data area style to a worksheet.

    Args:
        ws: Target worksheet.
        font_name: Font family name.
        col_count: Number of columns.
        row_count: Total rows including header.
    """
    body_font = _make_body_font(font_name)
    data_border = _make_data_border()
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_top = Alignment(horizontal="center", vertical="top")

    for row_idx in range(2, row_count + 1):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.border = data_border
            cell.alignment = wrap_alignment

    # Override alignment for specific columns that should be centered
    # (subclasses handle this per sheet type)


def apply_data_sheet_style(
    ws: Worksheet,
    font_name: str,
    segment_count: int,
) -> None:
    """Apply full styling to a data sheet (video segments).

    Args:
        ws: Data sheet worksheet.
        font_name: Font family name.
        segment_count: Number of data rows (segments).
    """
    total_rows = segment_count + 1  # +1 for header
    col_count = len(DATA_HEADERS)

    # Header style
    apply_header_style(ws, font_name, col_count)

    # Data area style
    apply_data_style(ws, font_name, col_count, total_rows)

    # Column-specific alignment
    center_top = Alignment(horizontal="center", vertical="top")
    for row_idx in range(2, total_rows + 1):
        # Index, Start, End = centered
        ws.cell(row=row_idx, column=1).alignment = center_top
        ws.cell(row=row_idx, column=2).alignment = center_top
        ws.cell(row=row_idx, column=3).alignment = center_top
        # English, Korean = left, wrap (already set by apply_data_style)

    # Column widths
    ws.column_dimensions["A"].width = COL_WIDTH_INDEX
    ws.column_dimensions["B"].width = COL_WIDTH_TIMESTAMP
    ws.column_dimensions["C"].width = COL_WIDTH_TIMESTAMP
    ws.column_dimensions["D"].width = COL_WIDTH_ENGLISH
    ws.column_dimensions["E"].width = COL_WIDTH_KOREAN

    # Freeze top row
    ws.freeze_panes = "A2"

    # Conditional formatting: highlight rows where Korean (col E) is empty
    if segment_count > 0:
        fail_fill = PatternFill(
            start_color=FAIL_ROW_BG, end_color=FAIL_ROW_BG, fill_type="solid"
        )
        range_str = f"A2:E{total_rows}"
        ws.conditional_formatting.add(
            range_str,
            FormulaRule(
                formula=[f'$E2=""'],
                fill=fail_fill,
            ),
        )


def apply_metadata_style(ws: Worksheet, font_name: str) -> None:
    """Apply styling to the _metadata sheet.

    Args:
        ws: _metadata worksheet.
        font_name: Font family name.
    """
    col_count = len(METADATA_HEADERS)
    total_rows = ws.max_row

    apply_header_style(ws, font_name, col_count)

    if total_rows > 1:
        apply_data_style(ws, font_name, col_count, total_rows)

    # Column widths
    for col_idx, width in enumerate(_METADATA_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Center alignment for specific columns
    center_cols = {5, 7, 8, 9, 10, 11, 13}  # duration, processed_at, counts, version
    center_top = Alignment(horizontal="center", vertical="top")
    for row_idx in range(2, total_rows + 1):
        for col_idx in center_cols:
            ws.cell(row=row_idx, column=col_idx).alignment = center_top

    # Freeze top row and auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Tab color
    ws.sheet_properties.tabColor = TAB_COLOR_METADATA


def apply_study_log_style(ws: Worksheet, font_name: str) -> None:
    """Apply styling to the _study_log sheet.

    Args:
        ws: _study_log worksheet.
        font_name: Font family name.
    """
    col_count = len(STUDY_LOG_HEADERS)
    total_rows = ws.max_row

    apply_header_style(ws, font_name, col_count)

    if total_rows > 1:
        apply_data_style(ws, font_name, col_count, total_rows)

    # Column widths
    for col_idx, width in enumerate(_STUDY_LOG_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row and auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Tab color
    ws.sheet_properties.tabColor = TAB_COLOR_STUDY_LOG

    # Conditional formatting: highlight "Not Started" status (col F)
    if total_rows > 1:
        not_started_fill = PatternFill(
            start_color=NOT_STARTED_BG, end_color=NOT_STARTED_BG, fill_type="solid"
        )
        status_range = f"F2:F{total_rows}"
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=['"Not Started"'],
                fill=not_started_fill,
            ),
        )


def apply_all_styles(wb: Workbook, font_name: str, data_sheet_name: str | None = None) -> None:
    """Apply styles to all relevant sheets in the workbook.

    Args:
        wb: The workbook.
        font_name: Font family name to use.
        data_sheet_name: Name of the newly created data sheet (if any).
    """
    # Style _metadata
    if METADATA_SHEET in wb.sheetnames:
        apply_metadata_style(wb[METADATA_SHEET], font_name)

    # Style _study_log
    if STUDY_LOG_SHEET in wb.sheetnames:
        apply_study_log_style(wb[STUDY_LOG_SHEET], font_name)

    # Style the data sheet
    if data_sheet_name and data_sheet_name in wb.sheetnames:
        ws = wb[data_sheet_name]
        segment_count = ws.max_row - 1  # subtract header
        apply_data_sheet_style(ws, font_name, segment_count)
